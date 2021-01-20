import xlrd
import xlwt
import openpyxl
from xlutils.copy import copy
from openpyxl import Workbook
from openpyxl import load_workbook
from xlwt import Workbook



def write_value_to_ExcelXLWT(path, sheetName, row, value):
    wb = Workbook(path)
    w_sheet = wb.get_sheet(sheetName)
    w_sheet.write(row, 2, value)
    wb.save(path)


def write_value_to_Excel(path, sheetName, row, value):
    rb = xlrd.open_workbook(path)
    wb = copy(rb)
    w_sheet = wb.get_sheet(sheetName)
    w_sheet.write(row, 2, value)
    wb.save(path)


def write_value_to_ExcelSheet(path, sheetName, row1, col, val):
    wb = load_workbook(path)
    ws = wb.get_sheet_by_name(sheetName)
    c1 = ws.cell(row=row1, column=col)
    c1.value = val
    wb.save(path)


def read_excelRow_file(path, sheetIndex, rowNumber):
    book = xlrd.open_workbook(path)
    first_sheet = book.sheet_by_index(int(sheetIndex))
    return first_sheet.row_values(int(rowNumber))


def get_total_rows(path, sheetIndex):
    book = xlrd.open_workbook(path)
    first_sheet = book.sheet_by_index(int(sheetIndex))
    return first_sheet.nrows


def read_excelCol_values(path, sheetIndex, colNumber):
    book = xlrd.open_workbook(path)
    first_sheet = book.sheet_by_index(int(sheetIndex))
    return first_sheet.col_values(int(colNumber))


def write_value_to_File(path, row1, col, val):
    wb = load_workbook(path)
    ws = wb['Sheet1']
    c1 = ws.cell(row=row1, column=col)
    c1.value = val
    wb.save(path)